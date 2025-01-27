import matplotlib
matplotlib.use('Agg')  # Use the 'Agg' backend to avoid Tkinter issues
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import io
import base64
from flask import Flask, render_template

app = Flask(__name__)

# Load the Excel file
data_path = "train_data.xlsx"

# Function to generate Total Cases plot
def plot_total_cases():
    wb = openpyxl.load_workbook(data_path)
    ws = wb.active
    lat, long, cases = [], [], []

    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=1).value and ws.cell(row=i, column=2).value and ws.cell(row=i, column=5).value:
            lat.append(float(ws.cell(row=i, column=1).value))
            long.append(float(ws.cell(row=i, column=2).value))
            cases.append(float(ws.cell(row=i, column=5).value))

    wb.close()

    x, y, z = np.array(lat), np.array(long), np.array(cases)

    # Plotting
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    ax.grid()
    ax.scatter(x, y, z, s=5)
    ax.set_title('Cases Graph From available info')
    ax.set_xlabel('Latitude', labelpad=20)
    ax.set_ylabel('Longitude', labelpad=20)
    ax.set_zlabel('Cases', labelpad=20)

    # Convert plot to PNG and then to base64 string
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode('utf8')
    plt.close(fig)
    
    return plot_url

# Function to generate Death Cases plot
def plot_death_cases():
    wb = openpyxl.load_workbook(data_path)
    ws = wb.active
    lat, long, deaths = [], [], []

    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=1).value and ws.cell(row=i, column=2).value and ws.cell(row=i, column=3).value:
            lat.append(float(ws.cell(row=i, column=1).value))
            long.append(float(ws.cell(row=i, column=2).value))
            deaths.append(float(ws.cell(row=i, column=3).value))

    wb.close()

    x, y, z = np.array(lat), np.array(long), np.array(deaths)

    # Plotting
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    ax.grid()
    ax.scatter(x, y, z, s=5)
    ax.set_title('Deaths Graph From available info')
    ax.set_xlabel('Latitude', labelpad=20)
    ax.set_ylabel('Longitude', labelpad=20)
    ax.set_zlabel('Deaths', labelpad=20)

    # Convert plot to PNG and then to base64 string
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode('utf8')
    plt.close(fig)
    
    return plot_url

# Landing page route
@app.route("/")
def home():
    return render_template("index.html")

# Route to show Total Cases graph
@app.route("/total_cases")
def total_cases():
    plot_url = plot_total_cases()
    return render_template("graph_page.html", plot_url=plot_url, title="Total Cases")

# Route to show Death Cases graph
@app.route("/death_cases")
def death_cases():
    plot_url = plot_death_cases()
    return render_template("graph_page.html", plot_url=plot_url, title="Death Cases")

if __name__ == "__main__":
    app.run(debug=True)
